import openpyxl
import json
from datetime import datetime

# Load the Excel file
wb = openpyxl.load_workbook('蟹江お散歩バス時刻表 Ver.2.xlsx')

timetable_data = {}

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    # Get headers (column names for bus numbers)
    headers = []
    for col in range(3, sheet.max_column + 1):
        cell_value = sheet.cell(1, col).value
        if cell_value:
            headers.append(str(cell_value))
    
    stops = []
    
    # Read each row (bus stop)
    for row in range(2, sheet.max_row + 1):
        stop_id = sheet.cell(row, 1).value
        stop_name = sheet.cell(row, 2).value
        
        if stop_id is None or stop_name is None:
            continue
        
        times = []
        for col in range(3, 3 + len(headers)):
            cell_value = sheet.cell(row, col).value
            
            # Convert datetime to time string
            if isinstance(cell_value, datetime):
                time_str = cell_value.strftime('%H:%M')
            elif cell_value:
                time_str = str(cell_value)
            else:
                time_str = '--'
            
            times.append(time_str)
        
        stops.append({
            'id': stop_id,
            'name': stop_name,
            'times': times
        })
    
    timetable_data[sheet_name] = {
        'bus_numbers': headers,
        'stops': stops
    }

# Save to JSON
with open('timetable_data.json', 'w', encoding='utf-8') as f:
    json.dump(timetable_data, f, ensure_ascii=False, indent=2)

print("Timetable data extracted successfully!")
print(f"Routes found: {list(timetable_data.keys())}")
for route, data in timetable_data.items():
    print(f"  {route}: {len(data['stops'])} stops, {len(data['bus_numbers'])} buses")
